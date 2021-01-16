# Installation of openpyxl in the "terminal": pip install openpyxl
# Installation of xlrd in the "terminal": pip install xlrd
# Installation of xlwt in the "terminal": pip install xlwt

# To use the code:
# 1- Navigate to the folder this file is in through the terminal: cd INSERT_PATH_TO_FOLDER
# 2- Run the .py file in the terminal: python sorting.py WORKBOOK1_NAME.xlsx WORKBOOK2_NAME.xlsx

#importing modules for manipulating csv, xlsx and xls files, and getting dates
import openpyxl, sys, xlrd, csv, xlwt
from datetime import datetime

# The file extenxions that are allowed in the program:
ALLOWED_EXTENSIONS = set(['xlsx','xls'])

# A function that returns True if the file is allowed in the program (has one of the allowed extensions) or not:
def allowed_file(filename):
	return '.' in filename and \
		   filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# Setting up (opening xlsx files needed) and checking if they are xls or xlsx
if len(sys.argv)== 3:
	if allowed_file(sys.argv[1]):
		success = False
		try:
			wb1 = xlrd.open_workbook(sys.argv[1])
			sheet = wb1.sheet_by_index(0)
			success = True
			xls1 = True
			maximum = sheet.nrows
		except:
			pass
		try:
			wb1 = openpyxl.load_workbook(sys.argv[1])
			sheet = wb1.get_sheet_by_name(wb1.sheetnames[0])
			success = True
			xls1 = False
			maximum = sheet.max_row+1
		except:
			pass
	else:
		print("\n\nCouldn't open the first workbook. The file attempted is "+sys.argv[1]+". The program cannot continue. (NOTE: Files must be .xlsx or .xls)\n\n")
		exit()

	if allowed_file(sys.argv[2]):
		success2 = False
		try:
			wb2 = xlrd.open_workbook(sys.argv[2])
			sheet_2 = wb2.sheet_by_index(0)
			success2 = True
			xls2 = True
			maximum2 = sheet_2.nrows
		except:
			pass
		try:
			wb2 = openpyxl.load_workbook(sys.argv[2])
			sheet_2 = wb2.get_sheet_by_name(wb2.sheetnames[0])
			success2 = True
			xls2 = False
			maximum2 = sheet_2.max_row+1
		except:
			pass
	else:
		print("\n\nCouldn't open the second workbook. The file attempted is "+sys.argv[2]+". The program cannot continue. (NOTE: Files must be .xlsx or .xls)\n\n")
		exit()

	if success ==False or success2 == False:
		print("\n\nCouldn't open one of the workbooks. The program cannot continue. (NOTE: Files don't exist)\n\n")
		exit()

	# Creating a new csv for first clinic and sets column names
	clinic1_wb = open('Clinic 1 '+str(datetime.date(datetime.now()))+'.csv','w')
	clinic1_writer = csv.writer(clinic1_wb)
	clinic1_writer.writerow(['First Name','Last Name','Email Address'])

	# Creating a new csv for second clinic and sets column names
	clinic2_wb = open('Clinic 2 '+str(datetime.date(datetime.now()))+'.csv','w')
	clinic2_writer = csv.writer(clinic2_wb)
	clinic2_writer.writerow(['First Name','Last Name','Email Address'])

	# List that helps with deleting
	empty_row = []

	# Searches the old xlsx/xls file and makes changes to the csv file based on what it finds
	for i in range(2, maximum2):

		# Getting all the information in WB2.xls
		if xls2 == False:
			clinic_name = sheet_2["A"+str(i)].value
			name = sheet_2["B"+str(i)].value
			checkin = sheet_2["C"+str(i)].value
		else:
			clinic_name = sheet_2.cell(i,0).value
			name = sheet_2.cell(i,1).value
			checkin = sheet_2.cell(i,2).value

		# Checks if check-in is empty and then if whole row is empty
		if checkin != None and checkin !=" " and checkin!=' ':
			if name != None and name != " " and name!=' ':
				try:

					# Splits the name column
					last_name, first_name = name.split("‚")

				# Trouble shooting
				except:
					print("Couldn't split "+name+" because it wasn't seperated by a comma")
					# Searches WB1.xls for the name. If found, it adds it to the clinic's workbook
				for k in range(2, maximum):
					if xls1 == False:
						last_name_2 = sheet["A"+str(k)].value.strip()
						first_name_2 = sheet["B"+str(k)].value.strip()
						email = sheet["C"+str(k)].value
					else:
						last_name_2 = sheet.cell(k,0).value.strip()
						first_name_2 = sheet.cell(k,1).value.strip()
						email = sheet.cell(k,2).value
					if last_name_2 == last_name.strip() and first_name_2 == first_name.strip():

						if clinic_name == "M":
							clinic1_writer.writerow([first_name,last_name,email])
						else:
							clinic2_writer.writerow([first_name,last_name,email])
						break
		else:
			empty_row.append(i)

	# Deletes rows with no check-in time
	if xls2 == False:
		empty_row.reverse()
		for i in empty_row:
			sheet_2.delete_rows(idx=i, amount=1)
		wb2.save(sys.argv[2])
	else:
		wb = xlwt.Workbook()
		sheet3 = wb.add_sheet('Sheet', cell_overwrite_ok=True) 
		for i in range(maximum2):
			if i not in empty_row:
				row = sheet_2.row_slice(i)
				rows_total = len(sheet3._Worksheet__rows)
				for column, heading in enumerate(row):
					sheet3.write(rows_total, column, heading.value)
		wb.save("resorted.xls")
		
	# Saves the changes made
	clinic1_wb.close()
	clinic2_wb.close()
	print("\n\n Everything was completed successfully\n\n")
else:
	print("\n\nPlease enter the running command in the correct format... sorting.py workbook1.xlsx workbook2.xlsx\n\n")